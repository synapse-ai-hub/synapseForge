import io
import logging
import os
import shutil
import subprocess
import tempfile
from dataclasses import dataclass
from typing import Optional, Tuple

from backend.agent.utils.error_logger import log_error

logger = logging.getLogger(__name__)

try:
    from docx import Document  # type: ignore
except ImportError as e:
    log_error(str(e), source="file_text_extractor.py")
    Document = None  # type: ignore

try:
    from pdfminer.high_level import extract_text as pdf_extract_text  # type: ignore
except ImportError as e:
    log_error(str(e), source="file_text_extractor.py")
    pdf_extract_text = None  # type: ignore

try:
    import pytesseract  # type: ignore
    from pytesseract.pytesseract import TesseractNotFoundError  # type: ignore
except ImportError as e:
    log_error(str(e), source="file_text_extractor.py")
    pytesseract = None  # type: ignore
    TesseractNotFoundError = None  # type: ignore

try:
    from pdf2image import convert_from_bytes  # type: ignore
    from pdf2image.exceptions import (  # type: ignore
        PDFInfoNotInstalledError,
        PDFPageCountError,
        PDFSyntaxError as PDF2ImageSyntaxError,
    )
except ImportError as e:
    log_error(str(e), source="file_text_extractor.py")
    convert_from_bytes = None  # type: ignore

try:
    from PIL import Image  # noqa: F401  # type: ignore
except ImportError as e:
    log_error(str(e), source="file_text_extractor.py")
    Image = None  # type: ignore

try:
    from openpyxl import load_workbook  # type: ignore
except ImportError as e:
    log_error(str(e), source="file_text_extractor.py")
    load_workbook = None  # type: ignore


class FileTextExtractionError(Exception):
    '''Base exception for errors that occur while extracting text from a file.'''


class UnsupportedFileTypeError(FileTextExtractionError):
    '''Raised when the file extension is not among the supported types.'''


class FileTooLargeError(FileTextExtractionError):
    '''Raised when the uploaded file exceeds the configured maximum size.'''


class MissingDependencyError(FileTextExtractionError):
    '''Raised when a required third-party library is not available.'''


SUPPORTED_TEXT_EXTENSIONS = {
    ".txt",
    ".md",
    ".csv",
    ".json",
    ".xml",
    ".yaml",
    ".yml",
    ".py",
}

ALLOWED_EXTENSIONS = SUPPORTED_TEXT_EXTENSIONS.union({".pdf", ".docx", ".doc", ".xlsx", ".xls"})

# Default limits - can be overridden via env vars if necessary.
MAX_UPLOAD_SIZE_MB = float(os.getenv("FILE_UPLOAD_MAX_MB", "50"))
MAX_OCR_PAGES = int(os.getenv("FILE_UPLOAD_OCR_MAX_PAGES", "10"))
PDF_TEXT_THRESHOLD = int(os.getenv("FILE_UPLOAD_PDF_TEXT_THRESHOLD", "80"))
OCR_DPI = int(os.getenv("FILE_UPLOAD_OCR_DPI", "200"))


@dataclass
class ExtractionResult:
    '''
    Container for the result of a text-extraction operation.

    Attributes
    ----------
    success : bool
        Indicates whether the extraction succeeded without raising an exception.
    text : str
        The extracted text if ``success`` is ``True``; otherwise an error placeholder.
    error_detail : Optional[str]
        Human-readable description of the error, if any.
    error_code : Optional[str]
        Machine-readable error identifier (e.g., ``file_too_large``).
    '''
    success: bool
    text: str
    error_detail: Optional[str] = None
    error_code: Optional[str] = None


def detect_extension(filename: str) -> str:
    '''
    Return the lower-case file extension of ``filename``.

    Parameters
    ----------
    filename : str
        Name of the file whose extension should be extracted.

    Returns
    -------
    str
        The extension, including the leading dot (e.g., ``".pdf"``).
    '''
    return os.path.splitext(filename)[1].lower()


def validate_extension(extension: str, filename: str) -> None:
    '''
    Verify that ``extension`` is allowed for upload.

    If the extension is not present in ``ALLOWED_EXTENSIONS``, an
    ``UnsupportedFileTypeError`` is raised with a message that lists the
    permitted extensions.

    Parameters
    ----------
    extension : str
        The file extension to validate (including the leading dot).
    filename : str
        The original filename, used only for constructing the error message.

    Raises
    ------
    UnsupportedFileTypeError
        If ``extension`` is not in ``ALLOWED_EXTENSIONS``.
    '''
    if extension not in ALLOWED_EXTENSIONS:
        raise UnsupportedFileTypeError(
            f"Formato no soportado para {filename}. Formatos permitidos: "
            f"{', '.join(sorted(ALLOWED_EXTENSIONS))}"
        )


def validate_size(file_size_bytes: int, filename: str) -> None:
    '''
    Ensure that the uploaded file does not exceed ``MAX_UPLOAD_SIZE_MB``.

    Parameters
    ----------
    file_size_bytes : int
        Size of the file in bytes.
    filename : str
        Name of the file, used for the error message.

    Raises
    ------
    FileTooLargeError
        If the file size exceeds the configured limit.
    '''
    max_bytes = int(MAX_UPLOAD_SIZE_MB * 1024 * 1024)
    if file_size_bytes > max_bytes:
        raise FileTooLargeError(
            f"{filename} excede el tamaño máximo de {MAX_UPLOAD_SIZE_MB} MB."
        )


def extract_text_from_bytes(filename: str, data: bytes) -> ExtractionResult:
    '''
    Extract plain text from a file represented by ``data`` according to its extension.

    The function determines the file type, validates the extension and size,
    and then dispatches to a format-specific extractor.  Supported formats
    include plain-text files, Markdown, CSV, JSON, XML, YAML, Python source,
    Microsoft Word (``.docx`` and ``.doc``), Excel (``.xlsx`` and ``.xls``),
    and PDF.  For PDFs, a fallback OCR step is performed when the extracted
    text is below ``PDF_TEXT_THRESHOLD`` characters.

    All validation and extraction errors are caught and transformed into an
    ``ExtractionResult`` with ``success=False`` and appropriate ``error_detail``
    and ``error_code`` fields.  Unexpected exceptions are logged and also
    result in ``success=False``.

    Parameters
    ----------
    filename : str
        Original name of the uploaded file (used for extension detection and
        error messages).
    data : bytes
        Raw file contents.

    Returns
    -------
    ExtractionResult
        An object describing whether extraction succeeded and containing the
        extracted text or an error placeholder.
    '''
    extension = detect_extension(filename)
    try:
        validate_extension(extension, filename)
        validate_size(len(data), filename)
        if extension in SUPPORTED_TEXT_EXTENSIONS:
            return ExtractionResult(True, _decode_text_bytes(data))
        if extension == ".docx":
            return ExtractionResult(True, _extract_docx_bytes(data))
        if extension == ".doc":
            return ExtractionResult(True, _extract_doc_bytes(data, filename))
        if extension == ".xlsx":
            return ExtractionResult(True, _extract_xlsx_bytes(data, filename))
        if extension == ".xls":
            return ExtractionResult(True, _extract_xls_bytes(data, filename))
        if extension == ".pdf":
            text = _extract_pdf_bytes(data, filename)
            return ExtractionResult(True, text)
        raise UnsupportedFileTypeError(
            f"Formato no soportado para {filename}."
        )
    except FileTooLargeError as exc:
        log_error(str(exc), source="file_text_extractor.py")
        logger.warning("Archivo demasiado grande %s: %s", filename, exc)
        return ExtractionResult(
            success=False,
            text=f"[Error al procesar {filename}: {exc}]",
            error_detail=str(exc),
            error_code="file_too_large",
        )
    except UnsupportedFileTypeError as exc:
        log_error(str(exc), source="file_text_extractor.py")
        logger.warning("Archivo no soportado %s: %s", filename, exc)
        return ExtractionResult(
            success=False,
            text=f"[Error al procesar {filename}: {exc}]",
            error_detail=str(exc),
            error_code="unsupported_type",
        )
    except MissingDependencyError as exc:
        log_error(str(exc), source="file_text_extractor.py")
        logger.warning("Faltan dependencias para %s: %s", filename, exc)
        return ExtractionResult(
            success=False,
            text=f"[Error al procesar {filename}: {exc}]",
            error_detail=str(exc),
            error_code="missing_dependency",
        )
    except FileTextExtractionError as exc:
        log_error(str(exc), source="file_text_extractor.py")
        logger.warning("Error procesando archivo %s: %s", filename, exc)
        return ExtractionResult(
            success=False,
            text=f"[Error al procesar {filename}: {exc}]",
            error_detail=str(exc),
            error_code="extraction_error",
        )
    except Exception as exc:  # pragma: no cover - safety net
        log_error(str(exc), source="file_text_extractor.py")
        logger.exception("Fallo inesperado procesando %s", filename)
        return ExtractionResult(
            success=False,
            text=f"[Error al procesar {filename}: fallo inesperado]",
            error_detail=str(exc),
            error_code="unexpected_error",
        )


def _decode_text_bytes(data: bytes) -> str:
    '''
    Decode a ``bytes`` object to ``str`` using UTF-8, falling back to Latin-1.

    Parameters
    ----------
    data : bytes
        Raw text data.

    Returns
    -------
    str
        Decoded text.

    Raises
    ------
    FileTextExtractionError
        If the data cannot be decoded with either UTF-8 or Latin-1.
    '''
    try:
        return data.decode("utf-8")
    except UnicodeDecodeError as e:
        log_error(str(e), source="file_text_extractor.py")
        try:
            return data.decode("latin-1")
        except UnicodeDecodeError as e:
            log_error(str(e), source="file_text_extractor.py")
            raise FileTextExtractionError(
                "No se logró decodificar el archivo como texto plano."
            )


def _extract_docx_bytes(data: bytes) -> str:
    '''
    Extract text from a ``.docx`` file stored in memory.

    The function uses ``python-docx`` to read paragraphs and tables.
    Paragraph texts are stripped of surrounding whitespace; table rows are
    concatenated with ``" | "`` separators.  Paragraph text appears before
    any extracted tables.

    Parameters
    ----------
    data : bytes
        Binary content of the ``.docx`` file.

    Returns
    -------
    str
        Combined plain-text representation of the document.

    Raises
    ------
    MissingDependencyError
        If ``python-docx`` is not installed.
    '''
    if Document is None:
        raise MissingDependencyError(
            "Dependencia python-docx no disponible en el servidor."
        )
    document = Document(io.BytesIO(data))
    paragraphs = [para.text.strip() for para in document.paragraphs]
    tables_text = _extract_docx_tables(document)
    combined = "\n".join([line for line in paragraphs if line])
    if tables_text:
        combined = f"{combined}\n{tables_text}".strip()
    return combined


def _extract_docx_tables(document) -> str:
    '''
    Extract table contents from a ``python-docx`` document.

    Each row is joined with ``" | "`` and rows are separated by newlines.

    Parameters
    ----------
    document : Document
        ``python-docx`` Document object.

    Returns
    -------
    str
        Text representation of all tables in the document.
    '''
    rows = []
    for table in getattr(document, "tables", []):
        for row in table.rows:
            cells = [cell.text.strip() for cell in row.cells]
            if any(cells):
                rows.append(" | ".join(cells))
    return "\n".join(rows)


def _extract_doc_bytes(data: bytes, filename: str) -> str:
    '''
    Convert a legacy ``.doc`` file to ``.docx`` using LibreOffice and extract text.

    The function locates the ``soffice`` executable, writes the input bytes to a
    temporary file, runs LibreOffice in headless mode to perform the conversion,
    and then delegates to ``_extract_docx_bytes``.  Errors during conversion
    raise ``FileTextExtractionError``.

    Parameters
    ----------
    data : bytes
        Binary content of the ``.doc`` file.
    filename : str
        Original filename (used for error messages).

    Returns
    -------
    str
        Extracted text from the converted document.

    Raises
    ------
    MissingDependencyError
        If LibreOffice ``soffice`` cannot be found.
    FileTextExtractionError
        If the conversion fails or the resulting ``.docx`` file is missing.
    '''
    soffice_path = _locate_soffice()
    if not soffice_path:
        raise MissingDependencyError(
            "Conversión DOC→DOCX no disponible (LibreOffice 'soffice' no encontrado)."
        )
    with tempfile.TemporaryDirectory() as tmpdir:
        input_path = os.path.join(tmpdir, "input.doc")
        output_path = os.path.join(tmpdir, "input.docx")
        with open(input_path, "wb") as tmp_file:
            tmp_file.write(data)
        try:
            result = subprocess.run(
                [soffice_path, "--headless", "--convert-to", "docx", input_path, "--outdir", tmpdir],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                check=False,
                timeout=60,
            )
        except subprocess.SubprocessError as exc:
            log_error(str(exc), source="file_text_extractor.py")
            raise FileTextExtractionError(
                f"No se pudo convertir {filename} a DOCX: {exc}"
            ) from exc

        if result.returncode != 0 or not os.path.exists(output_path):
            stderr = result.stderr.decode(errors="ignore")
            raise FileTextExtractionError(
                f"Conversión DOC→DOCX falló para {filename}: {stderr or 'sin detalles'}"
            )
        with open(output_path, "rb") as converted:
            return _extract_docx_bytes(converted.read())


def _extract_xlsx_bytes(data: bytes, filename: str) -> str:
    '''
    Extract text from an ``.xlsx`` workbook.

    The function writes the bytes to a temporary file, opens it with
    ``openpyxl`` in read-only mode, iterates over worksheets and rows,
    and concatenates cell values separated by tabs.  Worksheet titles are
    prefixed with ``[Hoja: <title>]``.  Empty rows are skipped.  The temporary
    file is removed after processing.

    Parameters
    ----------
    data : bytes
        Binary content of the ``.xlsx`` file.
    filename : str
        Original filename (used for error messages).

    Returns
    -------
    str
        Plain-text representation of the workbook.

    Raises
    ------
    MissingDependencyError
        If ``openpyxl`` is not installed.
    FileTextExtractionError
        If the workbook cannot be opened or contains no usable content.
    '''
    if load_workbook is None:
        raise MissingDependencyError(
            "Dependencia openpyxl no disponible en el servidor."
        )
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
        tmp_file.write(data)
        tmp_path = tmp_file.name
    try:
        wb = load_workbook(tmp_path, read_only=True, data_only=True)
    except Exception as exc:
        log_error(str(exc), source="file_text_extractor.py")
        raise FileTextExtractionError(
            f"No se pudo abrir el archivo Excel {filename}: {exc}"
        ) from exc
    lines = []
    try:
        for ws in wb.worksheets:
            lines.append(f"[Hoja: {ws.title}]")
            for row in ws.iter_rows(values_only=True):
                if not row:
                    continue
                values = ["" if cell is None else str(cell).strip() for cell in row]
                if any(values):
                    lines.append("\t".join(values))
    finally:
        try:
            wb.close()
        except Exception as e:
            log_error(str(e), source="file_text_extractor.py")
            pass
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception as e:
            log_error(str(e), source="file_text_extractor.py")
            pass

    text = "\n".join(lines).strip()
    if not text:
        raise FileTextExtractionError(
            f"No se encontró contenido utilizable en el archivo Excel {filename}."
        )
    return text


def _extract_xls_bytes(data: bytes, filename: str) -> str:
    '''
    Convert a legacy ``.xls`` file to ``.xlsx`` using LibreOffice and extract text.

    The conversion process mirrors ``_extract_doc_bytes``.  After conversion,
    the resulting ``.xlsx`` bytes are passed to ``_extract_xlsx_bytes`` for
    actual text extraction.

    Parameters
    ----------
    data : bytes
        Binary content of the ``.xls`` file.
    filename : str
        Original filename (used for error messages).

    Returns
    -------
    str
        Extracted text from the converted workbook.

    Raises
    ------
    MissingDependencyError
        If LibreOffice ``soffice`` cannot be found.
    FileTextExtractionError
        If the conversion fails or the resulting file is missing.
    '''
    soffice_path = _locate_soffice()
    if not soffice_path:
        raise MissingDependencyError(
            "Conversión XLS→XLSX no disponible (LibreOffice 'soffice' no encontrado)."
        )
    with tempfile.TemporaryDirectory() as tmpdir:
        input_path = os.path.join(tmpdir, "input.xls")
        output_path = os.path.join(tmpdir, "input.xlsx")
        with open(input_path, "wb") as tmp_file:
            tmp_file.write(data)
        try:
            result = subprocess.run(
                [soffice_path, "--headless", "--convert-to", "xlsx", input_path, "--outdir", tmpdir],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                check=False,
                timeout=60,
            )
        except subprocess.SubprocessError as exc:
            log_error(str(exc), source="file_text_extractor.py")
            raise FileTextExtractionError(
                f"No se pudo convertir {filename} a XLSX: {exc}"
            ) from exc

        if result.returncode != 0 or not os.path.exists(output_path):
            stderr = result.stderr.decode(errors="ignore")
            raise FileTextExtractionError(
                f"Conversión XLS→XLSX falló para {filename}: {stderr or 'sin detalles'}"
            )
        with open(output_path, "rb") as converted:
            return _extract_xlsx_bytes(converted.read(), filename)


def _locate_soffice() -> Optional[str]:
    '''
    Search the system PATH for the LibreOffice ``soffice`` executable.

    Returns
    -------
    Optional[str]
        Full path to ``soffice`` if found; otherwise ``None``.
    '''
    for candidate in ("soffice", "soffice.exe"):
        path = shutil.which(candidate)
        if path:
            return path
    return None


def _extract_pdf_bytes(data: bytes, filename: str) -> str:
    '''
    Extract text from a PDF file, falling back to OCR when necessary.

    The function first attempts to extract embedded text using
    ``pdfminer.six``.  If the resulting text length is below
    ``PDF_TEXT_THRESHOLD``, OCR is performed on the first
    ``MAX_OCR_PAGES`` pages using ``pdf2image`` and ``pytesseract``.
    The final result is the concatenation of the extracted text and any
    OCR output, with whitespace trimmed.

    Parameters
    ----------
    data : bytes
        Binary PDF content.
    filename : str
        Original filename (used for error messages).

    Returns
    -------
    str
        Extracted (and possibly OCR-augmented) text.

    Raises
    ------
    MissingDependencyError
        If ``pdfminer.six`` is not installed.
    FileTextExtractionError
        If PDF parsing fails.
    '''
    if pdf_extract_text is None:
        raise MissingDependencyError(
            "Dependencia pdfminer.six no disponible en el servidor."
        )
    try:
        text = pdf_extract_text(io.BytesIO(data))
    except Exception as exc:
        log_error(str(exc), source="file_text_extractor.py")
        raise FileTextExtractionError(
            f"No se pudo leer el PDF {filename}: {exc}"
        ) from exc

    cleaned = text.strip()
    if len(cleaned) >= PDF_TEXT_THRESHOLD:
        return cleaned

    # Try OCR fallback, but don't fail if poppler/OCR deps are missing
    try:
        ocr_text = _extract_pdf_via_ocr(data, filename)
        combined = "\n".join(filter(None, [cleaned, ocr_text]))
        return combined.strip() or cleaned
    except MissingDependencyError:
        # OCR dependencies (pdf2image, pytesseract, Pillow) not available
        logger.warning("OCR no disponible para %s (faltan dependencias), devolviendo texto extraído por pdfminer", filename)
        return cleaned
    except FileTextExtractionError as exc:
        # OCR failed (e.g., poppler not installed) - return what we have
        logger.warning("OCR falló para %s: %s, devolviendo texto extraído por pdfminer", filename, exc)
        return cleaned


def _extract_pdf_via_ocr(data: bytes, filename: str) -> str:
    '''
    Perform OCR on a PDF file using ``pdf2image`` and ``pytesseract``.

    The function renders up to ``MAX_OCR_PAGES`` pages at ``OCR_DPI`` DPI,
    runs Tesseract OCR (Spanish + English), and concatenates non-empty
    results separated by double newlines.  If OCR fails for a page, a
    warning is logged and processing continues.

    Parameters
    ----------
    data : bytes
        Binary PDF content.
    filename : str
        Original filename (used for error messages).

    Returns
    -------
    str
        OCR-derived text.

    Raises
    ------
    MissingDependencyError
        If any of ``pdf2image``, ``pytesseract`` or ``Pillow`` is missing,
        or if Tesseract executable is not found.
    FileTextExtractionError
        If PDF rendering fails or no text is extracted.
    '''
    if convert_from_bytes is None or pytesseract is None or Image is None:
        raise MissingDependencyError(
            "OCR no disponible: faltan dependencias (pdf2image, pytesseract o Pillow)."
        )
    try:
        images = convert_from_bytes(
            data,
            first_page=1,
            last_page=MAX_OCR_PAGES,
            dpi=OCR_DPI,
        )
    except (PDFInfoNotInstalledError, PDFPageCountError, PDF2ImageSyntaxError) as exc:
        log_error(str(exc), source="file_text_extractor.py")
        raise FileTextExtractionError(
            f"No se pudo preparar el PDF {filename} para OCR (poppler no instalado o error de PDF): {exc}"
        ) from exc
    except Exception as exc:  # pragma: no cover
        log_error(str(exc), source="file_text_extractor.py")
        raise FileTextExtractionError(
            f"Error inesperado preparando OCR para {filename}: {exc}"
        ) from exc

    ocr_results = []
    for index, image in enumerate(images, start=1):
        try:
            text = pytesseract.image_to_string(image, lang="spa+eng")
        except Exception as exc:  # pragma: no cover
            log_error(str(exc), source="file_text_extractor.py")
            if TesseractNotFoundError and isinstance(exc, TesseractNotFoundError):  # type: ignore[arg-type]
                raise MissingDependencyError(
                    "OCR no disponible: ejecutable de Tesseract no encontrado."
                ) from exc
            logger.warning("Error OCR en página %s de %s: %s", index, filename, exc)
            continue
        cleaned = text.strip()
        if cleaned:
            ocr_results.append(cleaned)

    if not ocr_results:
        raise FileTextExtractionError(
            f"No se extrajo texto mediante OCR para {filename}."
        )

    return "\n\n".join(ocr_results)


def is_extension_allowed(filename: str) -> Tuple[bool, str]:
    '''
    Determine whether a file's extension is permitted for upload.

    Parameters
    ----------
    filename : str
        Name of the file to check.

    Returns
    -------
    Tuple[bool, str]
        A tuple where the first element is ``True`` if the extension is in
        ``ALLOWED_EXTENSIONS`` and ``False`` otherwise; the second element is
        the detected lower-case extension.
    '''
    extension = detect_extension(filename)
    return extension in ALLOWED_EXTENSIONS, extension
